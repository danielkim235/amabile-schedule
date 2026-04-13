import os
from flask import Flask, render_template, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

app = Flask(__name__)

# 데이터베이스 설정 (PostgreSQL 또는 로컬 SQLite)
db_url = os.environ.get('DATABASE_URL')
if db_url and db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = db_url or 'sqlite:///schedule.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# 데이터베이스 모델 정의
class Schedule(db.Model):
    key = db.Column(db.String(100), primary_key=True)  # "2026-07-06_0_3층" 형태
    team_name = db.Column(db.String(200), nullable=False)

def parse_teams(file_path):
    teams = []
    if not os.path.exists(file_path):
        return teams
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    current_gen = ""
    for line in lines:
        line = line.strip()
        if not line: continue
        
        # Check for generation line (e.g., 41기)
        if line.endswith('기'):
            current_gen = line
            continue
            
        # Skip member lines
        if any(line.startswith(p) for p in ['V ', 'G ', 'B ', 'D ', 'K ', 'V:', 'G:', 'B:', 'D:', 'K:']):
            continue
        
        # Extract song title
        song_title = ""
        if '/' in line:
            song_title = line.split('/', 1)[1].strip()
        elif '-' in line:
            # Handle lines like "팀1 - 찬란" or "Vaundy- odoriko"
            parts = line.split('-', 1)
            # If the first part is just "팀X", the second part is the song
            if parts[0].strip().startswith('팀'):
                song_title = parts[1].strip()
            else:
                # For "Vaundy- odoriko", the part after '-' is the song
                song_title = parts[1].strip()
        
        if song_title:
            # Filter out blacklisted keywords
            blacklist = ['Oasis', '검정치마']
            if not any(k in song_title for k in blacklist):
                # Append generation info if available
                display_name = f"{song_title} ({current_gen})" if current_gen else song_title
                teams.append(display_name)
                
    return teams

@app.route('/')
def index():
    # 파일명을 영문(team_list.txt)으로 변경하여 서버 인식 오류 방지
    teams = parse_teams('./team_list.txt')
    start_date = datetime(2026, 7, 6)
    end_date = datetime(2026, 8, 23)
    weeks, current_week_start = [], start_date
    while current_week_start <= end_date:
        week = {'start': current_week_start, 'end': current_week_start + timedelta(days=6), 'days': []}
        for i in range(7):
            d = current_week_start + timedelta(days=i)
            week['days'].append({'date_str': d.strftime('%Y-%m-%d'), 'display': f"{d.strftime('%m/%d')} ({['월','화','수','목','금','토','일'][d.weekday()]})"})
        weeks.append(week)
        current_week_start += timedelta(days=7)

    time_slots = ["1타임 (10:00-12:00)", "2타임 (12:00-14:00)", "3타임 (14:00-16:00)", "4타임 (16:00-18:00)", "5타임 (18:00-20:00)", "6타임 (20:00-22:00)"]
    return render_template('index.html', teams=teams, weeks=weeks, time_slots=time_slots)

@app.route('/api/schedule', methods=['GET', 'POST'])
def schedule():
    if request.method == 'GET':
        items = Schedule.query.all()
        return jsonify({item.key: item.team_name for item in items})
    elif request.method == 'POST':
        data = request.json  # 전체 딕셔너리 전달됨
        # 전체 지우기 기능 대응을 위해 기존 데이터 삭제 후 일괄 삽입 (또는 업데이트 방식)
        Schedule.query.delete()
        for key, team in data.items():
            db.session.add(Schedule(key=key, team_name=team))
        db.session.commit()
        return jsonify({"status": "success"})

@app.route('/api/export')
def export_excel():
    items = Schedule.query.all()
    data = {item.key: item.team_name for item in items}
    
    start_date, end_date = datetime(2026, 7, 6), datetime(2026, 8, 23)
    time_slots = ["1타임 (10:00-12:00)", "2타임 (12:00-14:00)", "3타임 (14:00-16:00)", "4타임 (16:00-18:00)", "5타임 (18:00-20:00)", "6타임 (20:00-22:00)"]

    wb = Workbook()
    ws = wb.active
    ws.title = "합주 시간표"
    
    # (디자인 스타일 코드는 기존과 동일하므로 생략 없이 전체 포함)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    sub_header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    floor_3_label_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    floor_4_label_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold_font = Font(bold=True)

    current_week_start, week_num, curr_row = start_date, 1, 1
    while current_week_start <= end_date:
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=9)
        ws.cell(row=curr_row, column=1, value=f"WEEK {week_num}").fill = header_fill
        curr_row += 1
        
        ws.cell(row=curr_row, column=1, value="타임").fill = sub_header_fill
        ws.cell(row=curr_row, column=2, value="층").fill = sub_header_fill
        
        days_in_week = []
        for i in range(7):
            d = current_week_start + timedelta(days=i)
            days_in_week.append(d.strftime('%Y-%m-%d'))
            cell = ws.cell(row=curr_row, column=3+i, value=d.strftime('%m/%d'))
            cell.fill = sub_header_fill
        curr_row += 1
        
        for slot_idx, slot in enumerate(time_slots):
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row+1, end_column=1)
            ws.cell(row=curr_row, column=1, value=slot).alignment = center_alignment
            ws.cell(row=curr_row, column=2, value="3층").fill = floor_3_label_fill
            for i, d_str in enumerate(days_in_week):
                ws.cell(row=curr_row, column=3+i, value=data.get(f"{d_str}_{slot_idx}_3층", ""))
            curr_row += 1
            ws.cell(row=curr_row, column=2, value="4층").fill = floor_4_label_fill
            for i, d_str in enumerate(days_in_week):
                ws.cell(row=curr_row, column=3+i, value=data.get(f"{d_str}_{slot_idx}_4층", ""))
            curr_row += 1

        curr_row += 1
        current_week_start += timedelta(days=7)
        week_num += 1

    export_path = '합주_시간표_다운로드용.xlsx'
    wb.save(export_path)
    return send_file(export_path, as_attachment=True)

with app.app_context():
    db.create_all()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)